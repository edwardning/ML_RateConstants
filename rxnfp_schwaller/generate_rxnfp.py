import torch
import openpyxl
from rdkit import Chem
import numpy as np


def get_rxn_SMILES(rxn, spec_smiles):
    rxn = str(rxn)
    rxn = rxn.replace('<=>', '=')
    rxn = rxn.replace('=>', '=')
    reacts, prods = rxn.split('=')
    reacts = sorted(reacts.split('+'))  # sorted is necessary
    prods = sorted(prods.split('+'))  # sorted is necessary

    for s in reacts + prods:
        if s not in spec_smiles:
            print('Species lost: \t', s)
            return '***'

    rxn_smiles = ''
    for r in reacts:
        rxn_smiles += '{}.'.format(spec_smiles[r])
    rxn_smiles += '>>'
    for p in prods:
        rxn_smiles += '.{}'.format(spec_smiles[p])
    rxn_smiles = rxn_smiles.replace('.>>.', '>>')
    return rxn_smiles


def get_species_SMILES(sheet_obj):
    ans, temp = {}, {}
    rows = sheet_obj.iter_rows()
    for row in rows:
        if row[0].value in ['No.', None]:
            continue
        species = row[1].value.strip()
        SMILES = row[2].value.strip()
        InChi = row[3].value.strip()

        mol = Chem.inchi.MolFromInchi(InChi)
        canonical_SMILES = Chem.MolToSmiles(mol)

        if canonical_SMILES != SMILES:
            print('Different SMILES for {} ({}), {} {}'.format(species, InChi, SMILES, canonical_SMILES))
        if species in ans:
            print('Duplicate species: ', species)

        ans[species] = canonical_SMILES
    return ans


def parse_database(excel):
    """
    read database from raw excel
    generate reaction_smiles for each data items
    :param excel:
    :return:
    """
    wb = openpyxl.load_workbook(excel)
    smiles = get_species_SMILES(wb['species'])
    kinetics = wb['kinetics'].iter_rows()

    ans = {'rxn': [], 'rc': [], 'rxn_smiles': [], 'sub_mech': [], 'A': [], 'n': [], 'E': [], 'rxnfp': []}
    for row in kinetics:
        if row[0].value in ['No.', None]:
            continue
        if row[2].value == '!':
            continue
        ans['sub_mech'].append(str(row[1].value).strip())
        ans['rxn'].append(str(row[2].value).strip())
        ans['rc'].append(int(row[3].value))
        ans['A'].append(float(row[4].value))
        ans['n'].append(float(row[5].value))
        ans['E'].append(float(row[6].value))
        # ans['lnA_n_E'].append([math.log(ans['A'][-1]), ans['n'][-1], ans['E']])
        ans['rxn_smiles'].append(get_rxn_SMILES(ans['rxn'][-1], smiles))
    return ans


def generator(data=r"..\data\database.xlsx", model='default', save_as=r'..\data\database.npy'):
    print('Extract data from excel...')
    rxn_list = parse_database(data)

    print('Load rxnfp generator...')
    if model == 'ft':  # load rxnfp_generator from bert model fine_tuned on this database
        # Ref. to https://rxn4chemistry.github.io/rxnfp/fine_tune_bert_on_uspto_1k_tpl/
        from rxnfp_schwaller.transformer_fingerprints import (
            RXNBERTFingerprintGenerator, get_default_model_and_tokenizer, generate_fingerprints)
        model, tokenizer = get_default_model_and_tokenizer('bert_k_prediction')
        rxnfp_generator = RXNBERTFingerprintGenerator(model, tokenizer)
    else:  # load rxnfp_generator from default bert model by Schwaller
        from rxnfp.transformer_fingerprints import (
            RXNBERTFingerprintGenerator, get_default_model_and_tokenizer, generate_fingerprints)
        model, tokenizer = get_default_model_and_tokenizer()
        rxnfp_generator = RXNBERTFingerprintGenerator(model, tokenizer)

    print('Generating RxnFP...')
    for i, rxn_smiles in enumerate(rxn_list['rxn_smiles']):
        fp = rxnfp_generator.convert(rxn_smiles)
        rxn_list['rxnfp'].append(fp)

    print('Save to npy file...')
    for i in rxn_list:
        rxn_list[i] = np.array(rxn_list[i])
    torch.save(rxn_list, save_as)

    print('Enjoy!')
    return


if __name__ == '__main__':
    generator(data=r'..\data\database.xlsx', model='ft', save_as=r'..\data\database.npy')
